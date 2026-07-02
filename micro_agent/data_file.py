"""Run-scoped registration and bounded reading for data files."""

from __future__ import annotations

import csv
import hashlib
import json
import mimetypes
import secrets
import shutil
import time
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time as dt_time
from pathlib import Path
from typing import Any, Iterator

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024
MAX_READ_ROWS = 500
DEFAULT_READ_ROWS = 100
MAX_RETURN_CHARS = 50_000
MAX_SAMPLE_ROWS = 5
FILE_TTL_SECONDS = 3600

_FORMATS = {ext: ext[1:] for ext in (".csv", ".xls", ".xlsx", ".json", ".jsonl")}
_SUPPORTED_FILE_TYPES_HINT = "CSV（.csv）、Excel（.xls / .xlsx）、JSON（.json）、JSONL（.jsonl）"
_MIMES = {
    "csv": {"text/csv", "text/plain", "application/csv", "application/vnd.ms-excel"},
    "xls": {"application/vnd.ms-excel"},
    "xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    "json": {"application/json", "text/json", "text/plain"},
    "jsonl": {"application/jsonl", "application/x-ndjson", "application/json", "text/plain"},
}


class DataFileError(ValueError):
    def __init__(self, code: str, message: str, **details: Any) -> None:
        super().__init__(message)
        self.code, self.message, self.details = code, message, details

    def payload(self) -> dict[str, Any]:
        return {"error": {"code": self.code, "message": self.message, **self.details}}


@dataclass
class FileRecord:
    file_id: str
    original_name: str
    local_path: Path
    extension: str
    mime_type: str
    size_bytes: int
    sha256: str
    created_at: float
    expires_at: float
    run_id: str

    @property
    def format(self) -> str:
        return _FORMATS[self.extension]

    def public(self) -> dict[str, Any]:
        return {
            "file_id": self.file_id,
            "original_name": self.original_name,
            "format": self.format,
            "mime_type": self.mime_type,
            "size_bytes": self.size_bytes,
            "sha256": self.sha256,
            "created_at": self.created_at,
            "expires_at": self.expires_at,
        }


class FileRegistry:
    """Small in-memory index whose files live only for one run."""

    def __init__(self, root: Path, run_id: str) -> None:
        if not run_id or len(run_id) > 80 or any(c not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_-" for c in run_id):
            raise DataFileError("INVALID_RUN_ID", "run_id 格式无效。")
        self.root, self.run_id = root.resolve(), run_id
        self._records: dict[str, FileRecord] = {}

    async def register(self, upload: Any) -> FileRecord:
        name = Path(str(getattr(upload, "filename", "") or "").replace("\\", "/")).name[:255]
        name = "".join(char if char.isprintable() else "_" for char in name)
        ext = Path(name).suffix.lower()
        fmt = _FORMATS.get(ext)
        if not fmt:
            raise DataFileError(
                "UNSUPPORTED_FILE_TYPE",
                f"不支持的文件类型“{ext or '无扩展名'}”。请上传 {_SUPPORTED_FILE_TYPES_HINT}。",
            )
        mime = str(getattr(upload, "content_type", "") or mimetypes.guess_type(name)[0] or "application/octet-stream").split(";", 1)[0]
        if mime not in _MIMES[fmt] and mime != "application/octet-stream":
            raise DataFileError("FILE_TYPE_MISMATCH", "文件扩展名与 MIME 类型不匹配。")

        file_id = f"file_{secrets.token_hex(12)}"
        folder = self.root / self.run_id / file_id
        folder.mkdir(parents=True, exist_ok=False)
        path, digest, size = folder / f"original{ext}", hashlib.sha256(), 0
        try:
            with path.open("xb") as target:
                while chunk := await upload.read(1024 * 1024):
                    size += len(chunk)
                    if size > MAX_FILE_SIZE_BYTES:
                        raise DataFileError("FILE_TOO_LARGE", f"文件不能超过 {MAX_FILE_SIZE_BYTES} bytes。")
                    digest.update(chunk)
                    target.write(chunk)
            now = time.time()
            record = FileRecord(file_id, name, path, ext, mime, size, digest.hexdigest(), now, now + FILE_TTL_SECONDS, self.run_id)
            _validate_content(record)
            self._records[file_id] = record
            return record
        except DataFileError:
            shutil.rmtree(folder, ignore_errors=True)
            raise
        except Exception as exc:
            shutil.rmtree(folder, ignore_errors=True)
            raise DataFileError("FILE_REGISTER_FAILED", "文件注册失败。") from exc

    def get(self, file_id: str, run_id: str) -> FileRecord:
        record = self._records.get(str(file_id or ""))
        if not record:
            raise DataFileError("FILE_NOT_FOUND", f"文件“{file_id}”不存在。")
        if run_id != record.run_id:
            raise DataFileError("FILE_ACCESS_DENIED", "文件不属于当前运行。")
        if time.time() >= record.expires_at:
            raise DataFileError("FILE_EXPIRED", "文件已过期。")
        path, allowed = record.local_path, (self.root / record.run_id).resolve()
        try:
            if path.is_symlink() or not path.exists() or not path.resolve().is_relative_to(allowed):
                raise DataFileError("FILE_INVALID", "注册文件已被删除或路径无效。")
            if path.stat().st_size != record.size_bytes or _sha256(path) != record.sha256:
                raise DataFileError("FILE_TAMPERED", "注册文件在读取前发生了变化。")
        except OSError as exc:
            raise DataFileError("FILE_INVALID", "注册文件已被删除或无法读取。") from exc
        return record

    def metadata(self, file_ids: list[str]) -> list[dict[str, Any]]:
        return [self.get(file_id, self.run_id).public() for file_id in file_ids]

    def cleanup_expired(self) -> int:
        expired = [r for r in self._records.values() if time.time() >= r.expires_at]
        for record in expired:
            shutil.rmtree(record.local_path.parent, ignore_errors=True)
            self._records.pop(record.file_id, None)
        return len(expired)

    def cleanup(self) -> None:
        shutil.rmtree(self.root / self.run_id, ignore_errors=True)
        self._records.clear()


def inspect_data_file(record: FileRecord) -> dict[str, Any]:
    return _READERS[record.format][0](record)


def read_data_file(record: FileRecord, sheet: str | None, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    if record.format not in {"xls", "xlsx"} and sheet:
        raise DataFileError("SHEET_NOT_ALLOWED", "sheet 仅适用于 Excel 文件。")
    return _READERS[record.format][1](record, sheet, offset, limit, columns)


@contextmanager
def _open_csv(record: FileRecord) -> Iterator[tuple[Any, list[str]]]:
    encoding = _text_encoding(record.local_path)
    handle = record.local_path.open("r", encoding=encoding, newline="")
    warnings: list[str] = []
    try:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|") if sample else csv.excel
        except csv.Error:
            dialect = csv.excel
            warnings.append("未识别分隔符，按逗号处理。")
        csv.field_size_limit(1_000_000)
        yield csv.reader(handle, dialect), warnings
    except UnicodeDecodeError as exc:
        raise DataFileError("CSV_ENCODING_ERROR", "CSV 编码无法识别，请使用 UTF-8 或 GB18030。") from exc
    except csv.Error as exc:
        raise DataFileError("CSV_INVALID", "CSV 行或字段格式无效。") from exc
    finally:
        handle.close()


def _inspect_csv(record: FileRecord) -> dict[str, Any]:
    with _open_csv(record) as (reader, warnings):
        try:
            raw_headers = next(reader)
        except StopIteration:
            return _inspect_result(record, 0, [], [], ["CSV 文件为空。"])
        headers, header_warnings = _headers(raw_headers)
        samples, count = [], 0
        for values in reader:
            count += 1
            if len(samples) < MAX_SAMPLE_ROWS:
                samples.append(_row(headers, values))
    return _inspect_result(record, count, _typed_columns(headers, samples), samples, warnings + header_warnings)


def _read_csv(record: FileRecord, _sheet: str | None, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    with _open_csv(record) as (reader, warnings):
        try:
            headers, header_warnings = _headers(next(reader))
        except StopIteration:
            return _read_result(record, None, offset, limit, [], [], 0, ["CSV 文件为空。"])
        selected = _select_columns(headers, columns)
        rows, count = [], 0
        for values in reader:
            if offset <= count < offset + limit:
                row = _row(headers, values)
                rows.append({c: row.get(c) for c in selected})
            count += 1
    return _read_result(record, None, offset, limit, selected, rows, count, warnings + header_warnings)


class _JSONStream:
    def __init__(self, path: Path) -> None:
        self.file = path.open("r", encoding="utf-8-sig")
        self.buffer, self.pos, self.eof = "", 0, False
        self.decoder = json.JSONDecoder()

    def fill(self) -> None:
        self.buffer = self.buffer[self.pos:]
        self.pos = 0
        chunk = self.file.read(65_536)
        self.buffer += chunk
        self.eof = not chunk

    def peek(self) -> str:
        while True:
            while self.pos < len(self.buffer) and self.buffer[self.pos].isspace():
                self.pos += 1
            if self.pos < len(self.buffer):
                return self.buffer[self.pos]
            if self.eof:
                return ""
            self.fill()

    def take(self, expected: str) -> None:
        if self.peek() != expected:
            raise DataFileError("JSON_INVALID", "JSON 文件格式错误。")
        self.pos += 1

    def value(self) -> Any:
        while True:
            self.peek()
            try:
                value, self.pos = self.decoder.raw_decode(self.buffer, self.pos)
                return value
            except json.JSONDecodeError as exc:
                if self.eof:
                    raise DataFileError("JSON_INVALID", "JSON 文件格式错误。") from exc
                self.fill()

    def finish(self) -> None:
        if self.peek():
            raise DataFileError("JSON_INVALID", "JSON 文件存在多余内容。")


def _json_array(stream: _JSONStream) -> Iterator[dict[str, Any]]:
    stream.take("[")
    if stream.peek() == "]":
        stream.take("]")
        return
    while True:
        value = stream.value()
        yield value if isinstance(value, dict) else {"value": value}
        separator = stream.peek()
        if separator == "]":
            stream.take("]")
            return
        stream.take(",")


def _iter_json(record: FileRecord, state: dict[str, Any]) -> Iterator[dict[str, Any]]:
    stream = _JSONStream(record.local_path)
    try:
        first = stream.peek()
        if first == "[":
            yield from _json_array(stream)
        elif first == "{":
            stream.take("{")
            pairs, has_data = [], False
            if stream.peek() != "}":
                while True:
                    key = stream.value()
                    if not isinstance(key, str):
                        raise DataFileError("JSON_INVALID", "JSON 对象键必须是字符串。")
                    stream.take(":")
                    if key == "data" and stream.peek() == "[":
                        has_data = True
                        yield from _json_array(stream)
                    else:
                        pairs.append({"key": key, "value": _brief(stream.value())})
                    if stream.peek() == "}":
                        break
                    stream.take(",")
            stream.take("}")
            if not has_data:
                state["warning"] = "JSON 不是表格数组，按顶层键值读取。"
                yield from pairs
        elif first:
            state["warning"] = "JSON 顶层不是数组或对象。"
            yield {"value": stream.value()}
        else:
            raise DataFileError("JSON_INVALID", "JSON 文件为空。")
        stream.finish()
    except UnicodeDecodeError as exc:
        raise DataFileError("JSON_INVALID", "JSON 必须使用 UTF-8 编码。") from exc
    finally:
        stream.file.close()


def _inspect_json(record: FileRecord) -> dict[str, Any]:
    state, samples, headers, count = {}, [], [], 0
    for row in _iter_json(record, state):
        count += 1
        for key in row:
            if key not in headers:
                headers.append(key)
        if len(samples) < MAX_SAMPLE_ROWS:
            samples.append(row)
    warnings = [state["warning"]] if state.get("warning") else []
    return _inspect_result(record, count, _typed_columns(headers, samples), samples, warnings)


def _read_json(record: FileRecord, _sheet: str | None, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    state, rows, headers, count = {}, [], [], 0
    for row in _iter_json(record, state):
        for key in row:
            if key not in headers:
                headers.append(key)
        if offset <= count < offset + limit:
            rows.append(row)
        count += 1
    selected = _select_columns(headers, columns)
    page = [{c: row.get(c) for c in selected} for row in rows]
    warnings = [state["warning"]] if state.get("warning") else []
    return _read_result(record, None, offset, limit, selected, page, count, warnings)


def _iter_jsonl(record: FileRecord) -> Iterator[dict[str, Any]]:
    try:
        with record.local_path.open("r", encoding="utf-8-sig") as handle:
            for line_no, line in enumerate(handle, 1):
                if not line.strip():
                    continue
                try:
                    value = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise DataFileError("JSONL_INVALID", f"JSONL 第 {line_no} 行格式错误。", line=line_no) from exc
                if not isinstance(value, dict):
                    raise DataFileError("JSONL_INVALID", f"JSONL 第 {line_no} 行不是对象。", line=line_no)
                yield value
    except UnicodeDecodeError as exc:
        raise DataFileError("JSONL_ENCODING_ERROR", "JSONL 必须使用 UTF-8 编码。") from exc


def _inspect_jsonl(record: FileRecord) -> dict[str, Any]:
    samples, headers, count = [], [], 0
    for row in _iter_jsonl(record):
        count += 1
        for key in row:
            if key not in headers:
                headers.append(key)
        if len(samples) < MAX_SAMPLE_ROWS:
            samples.append(row)
    return _inspect_result(record, count, _typed_columns(headers, samples), samples, ["JSONL 文件为空。"] if not count else [])


def _read_jsonl(record: FileRecord, _sheet: str | None, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    rows, headers, count = [], [], 0
    for row in _iter_jsonl(record):
        for key in row:
            if key not in headers:
                headers.append(key)
        if offset <= count < offset + limit:
            rows.append(row)
        count += 1
    selected = _select_columns(headers, columns)
    page = [{c: row.get(c) for c in selected} for row in rows]
    return _read_result(record, None, offset, limit, selected, page, count, [])


def _workbook(record: FileRecord) -> Any:
    try:
        import openpyxl
        return openpyxl.load_workbook(record.local_path, read_only=True, data_only=True)
    except Exception as exc:
        raise DataFileError("XLSX_INVALID", "XLSX 文件无法打开。") from exc


def _inspect_xlsx(record: FileRecord) -> dict[str, Any]:
    wb, sheets = _workbook(record), []
    try:
        for ws in wb.worksheets:
            iterator = ws.iter_rows(values_only=True)
            raw_headers = next(iterator, None)
            if raw_headers is None:
                sheets.append({"name": ws.title, "row_count": 0, "column_count": 0, "headers": [], "sample_rows": [], "warnings": ["工作表为空。"]})
                continue
            sheets.append(_inspect_sheet(ws.title, raw_headers, ws.max_column, iterator))
    finally:
        wb.close()
    return {"file_id": record.file_id, "format": "xlsx", "size_bytes": record.size_bytes, "sheets": sheets, "warnings": []}


def _read_xlsx(record: FileRecord, sheet: str | None, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    wb = _workbook(record)
    try:
        if sheet and sheet not in wb.sheetnames:
            raise DataFileError("SHEET_NOT_FOUND", f"工作表“{sheet}”不存在。", available_sheets=wb.sheetnames)
        ws = wb[sheet] if sheet else wb.active
        iterator = ws.iter_rows(values_only=True)
        raw_headers = next(iterator, None)
        if raw_headers is None:
            return _read_result(record, ws.title, offset, limit, [], [], 0, ["工作表为空。"])
        return _read_sheet(record, ws.title, raw_headers, ws.max_column, iterator, offset, limit, columns)
    finally:
        wb.close()


def _xls(record: FileRecord) -> Any:
    try:
        import xlrd
        return xlrd.open_workbook(record.local_path, on_demand=True)
    except Exception as exc:
        raise DataFileError("XLS_INVALID", "XLS 文件无法打开。") from exc


def _xls_values(book: Any, sheet: Any, row: int) -> list[Any]:
    import xlrd
    values = []
    for cell in sheet.row(row):
        value = cell.value
        if cell.ctype == xlrd.XL_CELL_DATE:
            value = xlrd.xldate_as_datetime(value, book.datemode)
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            value = bool(value)
        elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
            value = None
        values.append(value)
    return values


def _inspect_xls(record: FileRecord) -> dict[str, Any]:
    book, sheets = _xls(record), []
    try:
        for sheet in book.sheets():
            if not sheet.nrows:
                sheets.append({"name": sheet.name, "row_count": 0, "column_count": 0, "headers": [], "sample_rows": [], "warnings": ["工作表为空。"]})
                continue
            rows = (_xls_values(book, sheet, row) for row in range(1, sheet.nrows))
            sheets.append(_inspect_sheet(sheet.name, _xls_values(book, sheet, 0), sheet.ncols, rows))
    finally:
        book.release_resources()
    return {"file_id": record.file_id, "format": "xls", "size_bytes": record.size_bytes, "sheets": sheets, "warnings": []}


def _read_xls(record: FileRecord, sheet_name: str | None, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    book = _xls(record)
    try:
        if sheet_name and sheet_name not in book.sheet_names():
            raise DataFileError("SHEET_NOT_FOUND", f"工作表“{sheet_name}”不存在。", available_sheets=book.sheet_names())
        sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
        if not sheet.nrows:
            return _read_result(record, sheet.name, offset, limit, [], [], 0, ["工作表为空。"])
        rows = (_xls_values(book, sheet, row) for row in range(1, sheet.nrows))
        return _read_sheet(record, sheet.name, _xls_values(book, sheet, 0), sheet.ncols, rows, offset, limit, columns)
    finally:
        book.release_resources()


def _inspect_sheet(name: str, raw_headers: Any, width: int, rows: Any) -> dict[str, Any]:
    headers, warnings = _headers(raw_headers, width)
    samples, count = [], 0
    for values in rows:
        if not any(value not in (None, "") for value in values):
            continue
        count += 1
        if len(samples) < MAX_SAMPLE_ROWS:
            samples.append(_row(headers, values))
    return {"name": name, "row_count": count, "column_count": len(headers), "headers": headers, "columns": _typed_columns(headers, samples), "sample_rows": samples, "warnings": warnings}


def _read_sheet(record: FileRecord, name: str, raw_headers: Any, width: int, source: Any, offset: int, limit: int, columns: list[str] | None) -> dict[str, Any]:
    headers, warnings = _headers(raw_headers, width)
    selected, rows, count = _select_columns(headers, columns), [], 0
    for values in source:
        if not any(value not in (None, "") for value in values):
            continue
        if offset <= count < offset + limit:
            row = _row(headers, values)
            rows.append({column: row.get(column) for column in selected})
        count += 1
    return _read_result(record, name, offset, limit, selected, rows, count, warnings)


def _inspect_result(record: FileRecord, count: int, columns: list[Any], samples: list[Any], warnings: list[str]) -> dict[str, Any]:
    return {"file_id": record.file_id, "format": record.format, "size_bytes": record.size_bytes, "row_count": count, "columns": columns, "sample_rows": samples, "warnings": warnings}


def _read_result(record: FileRecord, sheet: str | None, offset: int, limit: int, columns: list[str], rows: list[Any], total: int, warnings: list[str]) -> dict[str, Any]:
    return {"file_id": record.file_id, "format": record.format, "sheet": sheet, "offset": offset, "requested_limit": limit, "returned_rows": len(rows), "total_rows": total, "columns": columns, "rows": rows, "truncated": False, "warnings": warnings}


def _headers(values: Any, width: int | None = None) -> tuple[list[str], list[str]]:
    values, warnings, seen = list(values or []), [], {}
    width = max(width or 0, len(values))
    headers = []
    for index in range(width):
        raw = values[index] if index < len(values) else None
        base = str(raw).strip() if raw is not None else ""
        if not base:
            base = f"column_{index + 1}"
            warnings.append("存在空表头，已生成稳定列名。")
        seen[base] = seen.get(base, 0) + 1
        name = base if seen[base] == 1 else f"{base}_{seen[base]}"
        if name != base:
            warnings.append(f"重复表头“{base}”已重命名为“{name}”。")
        headers.append(name)
    return headers, list(dict.fromkeys(warnings))


def _row(headers: list[str], values: Any) -> dict[str, Any]:
    values = list(values or [])
    return {name: _json_value(values[i]) if i < len(values) else None for i, name in enumerate(headers)}


def _select_columns(available: list[str], requested: list[str] | None) -> list[str]:
    if requested is None:
        return available
    missing = [name for name in requested if name not in available]
    if missing:
        raise DataFileError("COLUMN_NOT_FOUND", f"列不存在：{', '.join(missing)}。", available_columns=available)
    return requested


def _typed_columns(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [{"name": name, "inferred_type": _inferred([row.get(name) for row in rows])} for name in headers]


def _inferred(values: list[Any]) -> str:
    kinds = {_kind(value) for value in values}
    non_null = kinds - {"null"}
    return "null" if not non_null else next(iter(non_null)) if len(non_null) == 1 else "mixed"


def _kind(value: Any) -> str:
    if value is None or value == "":
        return "null"
    if isinstance(value, bool) or str(value).strip().lower() in {"true", "false"}:
        return "boolean"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    text = str(value).strip()
    try:
        float(text.replace(",", ""))
        return "number"
    except ValueError:
        pass
    try:
        datetime.fromisoformat(text.replace("Z", "+00:00"))
        return "datetime"
    except ValueError:
        return "string"


def _json_value(value: Any) -> Any:
    return value.isoformat() if isinstance(value, (datetime, date, dt_time)) else value


def _brief(value: Any) -> Any:
    if not isinstance(value, (dict, list)):
        return _json_value(value)
    text = dump_result(value)
    return text if len(text) <= 1000 else text[:1000] + "…"


def bounded_result(payload: dict[str, Any]) -> dict[str, Any]:
    if len(dump_result(payload)) <= MAX_RETURN_CHARS:
        return payload
    payload["truncated"] = True
    payload.setdefault("warnings", []).append(f"结果超过 {MAX_RETURN_CHARS} 字符，已截断。")
    lists = [payload.get("rows"), payload.get("sample_rows")]
    lists += [sheet.get("sample_rows") for sheet in payload.get("sheets", [])]
    for rows in [r for r in lists if isinstance(r, list)]:
        while rows and len(dump_result(payload)) > MAX_RETURN_CHARS:
            rows.pop()
    if "returned_rows" in payload:
        payload["returned_rows"] = len(payload.get("rows") or [])
    if len(dump_result(payload)) > MAX_RETURN_CHARS:
        return {"file_id": payload.get("file_id"), "format": payload.get("format"), "truncated": True, "warnings": ["文件结构元数据超过返回限制。"]}
    return payload


def _validate_content(record: FileRecord) -> None:
    if record.format == "xls":
        if _head(record.local_path, 8) != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            raise DataFileError("FILE_TYPE_MISMATCH", "文件不是有效的 XLS。")
        return
    if record.format == "xlsx":
        if not zipfile.is_zipfile(record.local_path):
            raise DataFileError("FILE_TYPE_MISMATCH", "文件不是有效的 XLSX。")
        with zipfile.ZipFile(record.local_path) as archive:
            if not {"[Content_Types].xml", "xl/workbook.xml"}.issubset(archive.namelist()):
                raise DataFileError("FILE_TYPE_MISMATCH", "文件不是有效的 XLSX。")
        return
    raw = _head(record.local_path)
    if record.format == "csv":
        _decode_text(raw, "CSV_ENCODING_ERROR")
    elif record.format == "json":
        text = _decode_text(raw, "JSON_INVALID").lstrip("\ufeff \t\r\n")
        if text and text[0] not in "[{":
            raise DataFileError("FILE_TYPE_MISMATCH", "文件内容不是 JSON。")
    elif record.format == "jsonl":
        text = _decode_text(raw, "JSONL_INVALID")
        first = next((line for line in text.splitlines() if line.strip()), "")
        if first:
            try:
                if not isinstance(json.loads(first), dict):
                    raise ValueError
            except (json.JSONDecodeError, ValueError) as exc:
                raise DataFileError("FILE_TYPE_MISMATCH", "文件内容不是 JSONL 对象流。") from exc


def _decode_text(raw: bytes, code: str) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise DataFileError(code, "文件编码无法识别。")


def _text_encoding(path: Path) -> str:
    raw = _head(path)
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            raw.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise DataFileError("CSV_ENCODING_ERROR", "CSV 编码无法识别，请使用 UTF-8 或 GB18030。")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _head(path: Path, size: int = 65_536) -> bytes:
    with path.open("rb") as handle:
        return handle.read(size)


def dump_result(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":"))


_READERS = {
    "csv": (_inspect_csv, _read_csv), "xls": (_inspect_xls, _read_xls),
    "xlsx": (_inspect_xlsx, _read_xlsx), "json": (_inspect_json, _read_json),
    "jsonl": (_inspect_jsonl, _read_jsonl),
}


__all__ = [
    "FileRegistry", "FileRecord", "DataFileError", "inspect_data_file", "read_data_file",
    "bounded_result", "dump_result",
    "MAX_FILE_SIZE_BYTES", "MAX_READ_ROWS", "DEFAULT_READ_ROWS", "MAX_RETURN_CHARS",
    "MAX_SAMPLE_ROWS", "FILE_TTL_SECONDS",
]
